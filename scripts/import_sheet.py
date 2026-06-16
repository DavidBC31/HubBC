#!/usr/bin/env python3
"""
Import + normalisation du `pointage suivi.xlsx` -> data/relances.json

Toute la logique "sale" (texte libre -> champs explicites) vit ICI, une fois.
Le front Next.js ne consomme que le JSON normalisé.

Usage: python3 scripts/import_sheet.py [chemin_xlsx]
"""
import sys, json, os, datetime
import openpyxl

XLSX = sys.argv[1] if len(sys.argv) > 1 else "/tmp/pointage_suivi.xlsx"
OUT = os.path.join(os.path.dirname(__file__), "..", "data", "relances.json")

SKIP_KEYWORDS = ["LIAISON PIMS", "AUTONOME", "AUTOMATIQUE", "COMPLET",
                 "NE PAS RELANCER", "NE DONNE PAS", "NE VEUT PAS", "NE NOUS LE DONNE"]
CADENCE_JOURS = {"LUNDI": ["LUNDI"], "MARDI": ["MARDI"], "MERCREDI": ["MERCREDI"],
                 "MERCRDI": ["MERCREDI"], "JEUDI": ["JEUDI"], "VENDREDI": ["VENDREDI"]}


def cell(r, i): return r[i] if i < len(r) and r[i] is not None else None
def txt(v):
    if v is None or isinstance(v, (datetime.datetime, datetime.date)): return ""
    return str(v).strip()
def as_iso(v):
    if isinstance(v, datetime.datetime): return v.date().isoformat()
    if isinstance(v, datetime.date): return v.isoformat()
    return None


def classify(commentaire, relance_raw):
    blob = f"{txt(commentaire)} {txt(relance_raw)}".upper()
    for kw in SKIP_KEYWORDS:
        if kw in blob:
            return "SKIP", [], kw
    if "MEV" in blob:
        return "SKIP", [], "MEV (pas encore en vente)"
    jours = []
    for kw, js in CADENCE_JOURS.items():
        if kw in blob:
            for j in js:
                if j not in jours: jours.append(j)
    if ("SEMAINE" in blob or "HEBDO" in blob) and not jours:
        jours = ["LUNDI"]
    if "MOIS" in blob:
        return "MENSUEL", [], "1 fois / mois"
    if not jours:
        jours = ["LUNDI"]  # défaut cahier des charges
    return "RELANCE", jours, ""


def parse_grouped(ws, source):
    out, artist = [], None
    for r in ws.iter_rows(min_row=2, values_only=True):
        manif, ville = cell(r, 0), cell(r, 1)
        if manif and not ville:
            artist = str(manif).strip(); continue
        if not any(r) or not ville: continue
        action, cad, raison = classify(cell(r, 6), cell(r, 7))
        out.append({
            "source": source, "artiste": artist or txt(manif),
            "ville": txt(ville), "salle": txt(cell(r, 2)),
            "date_concert": as_iso(cell(r, 3)), "mail": txt(cell(r, 4)),
            "commentaire": txt(cell(r, 6)),
            "relance_raw": txt(cell(r, 7)) or as_iso(cell(r, 7)) or "",
            "dernier_recu": as_iso(cell(r, 8)),
            "identifiant": txt(cell(r, 10)), "mdp": txt(cell(r, 11)),
            "action": action, "cadence": cad, "raison_skip": raison,
            "billet_tiers": False, "spectacles": "",
        })
    return out


def parse_billets_tiers(ws):
    out = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        nom = cell(r, 0)
        if not nom or not str(nom).strip(): continue
        if str(nom).strip().upper().startswith("ARCHIVE"): break
        action, cad, raison = classify(cell(r, 5), cell(r, 3))
        out.append({
            "source": "BILLETS TIERS", "artiste": str(nom).strip(),
            "ville": "", "salle": "", "date_concert": None,
            "mail": txt(cell(r, 1)), "commentaire": txt(cell(r, 5)),
            "relance_raw": txt(cell(r, 3)), "dernier_recu": as_iso(cell(r, 4)),
            "identifiant": txt(cell(r, 6)), "mdp": txt(cell(r, 7)),
            "action": action, "cadence": cad, "raison_skip": raison,
            "billet_tiers": True, "spectacles": txt(cell(r, 2)),
        })
    return out


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    entries = []
    entries += parse_grouped(wb["RELANCES  PIMS A SAISIR"], "RELANCES")
    entries += parse_grouped(wb["AUTONOMES  PIMS A SAISIR"], "AUTONOMES")
    entries += parse_billets_tiers(wb["BILLETS TIERS"])
    for i, e in enumerate(entries):
        e["id"] = i + 1
    payload = {
        "generated_at": datetime.datetime.now().isoformat(timespec="seconds"),
        "source_file": os.path.basename(XLSX),
        "count": len(entries),
        "entries": entries,
    }
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    print(f"OK -> {os.path.relpath(OUT)} ({len(entries)} entrées)")


if __name__ == "__main__":
    main()
