#!/usr/bin/env python3
"""
Moteur de relance — DRY RUN (ne touche à rien : ni Gmail, ni PIMS).

Lit l'export `pointage suivi.xlsx`, normalise les onglets RELANCES / AUTONOMES /
BILLETS TIERS en entrées propres, applique la logique d'aiguillage, et génère
le lot de relances pour une date donnée (par défaut : aujourd'hui).

But : valider l'interprétation des données + le template, avant de brancher
la génération de brouillons Gmail.

Usage:
    python3 scripts/relance_preview.py [chemin_xlsx] [YYYY-MM-DD]
"""
import sys, re, datetime, collections
import openpyxl

XLSX = sys.argv[1] if len(sys.argv) > 1 else "/tmp/pointage_suivi.xlsx"
RUN_DATE = (datetime.date.fromisoformat(sys.argv[2])
            if len(sys.argv) > 2 else datetime.date.today())

JOURS = {0: "LUNDI", 1: "MARDI", 2: "MERCREDI", 3: "JEUDI",
         4: "VENDREDI", 5: "SAMEDI", 6: "DIMANCHE"}
JOUR_RUN = JOURS[RUN_DATE.weekday()]

# --- Mots-clés d'aiguillage (cherchés dans COMMENTAIRE + RELANCE, en MAJ) ---
SKIP_KEYWORDS = [
    "LIAISON PIMS", "AUTONOME", "AUTOMATIQUE", "COMPLET",
    "NE PAS RELANCER", "NE DONNE PAS", "NE VEUT PAS", "NE NOUS LE DONNE",
]
# Cadences explicites -> jours où la relance part
CADENCE_JOURS = {
    "LUNDI": {"LUNDI"}, "MARDI": {"MARDI"}, "MERCREDI": {"MERCREDI"},
    "MERCRDI": {"MERCREDI"}, "JEUDI": {"JEUDI"}, "VENDREDI": {"VENDREDI"},
}


def cell(row, i):
    return row[i] if i < len(row) and row[i] is not None else None


def txt(v):
    if v is None:
        return ""
    if isinstance(v, (datetime.datetime, datetime.date)):
        return ""  # une date n'est pas un statut texte
    return str(v).strip()


def as_date(v):
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    return None


def classify(commentaire, relance_raw):
    """Retourne (action, cadence_jours, raison)."""
    blob = f"{txt(commentaire)} {txt(relance_raw)}".upper()
    for kw in SKIP_KEYWORDS:
        if kw in blob:
            return ("SKIP", set(), kw)
    if "MEV" in blob:  # pas encore en vente / mise en vente à venir
        return ("SKIP", set(), "MEV (pas encore en vente)")
    # cadence spécifique ?
    jours = set()
    for kw, js in CADENCE_JOURS.items():
        if kw in blob:
            jours |= js
    if "SEMAINE" in blob or "HEBDO" in blob:
        jours |= {"LUNDI"}  # 1x/semaine -> lundi par défaut
    if "MOIS" in blob:
        return ("MENSUEL", set(), "1 fois / mois")
    if not jours:
        jours = {"LUNDI"}  # défaut cahier des charges : relance le lundi
    return ("RELANCE", jours, "")


def parse_relances(ws, source):
    """Onglet groupé par artiste (ligne section = MANIF rempli, VILLE vide)."""
    entries = []
    current_artist = None
    for r in ws.iter_rows(min_row=2, values_only=True):
        manif, ville = cell(r, 0), cell(r, 1)
        if manif and not ville:
            current_artist = str(manif).strip()
            continue
        if not any(r):
            continue
        if not ville:
            continue
        action, jours, raison = classify(cell(r, 6), cell(r, 7))
        entries.append({
            "source": source,
            "artiste": current_artist or txt(manif),
            "ville": txt(ville),
            "salle": txt(cell(r, 2)),
            "date_concert": as_date(cell(r, 3)),
            "mail": txt(cell(r, 4)),
            "commentaire": txt(cell(r, 6)),
            "relance_raw": txt(cell(r, 7)) or _fmt(as_date(cell(r, 7))),
            "dernier_recu": as_date(cell(r, 8)),
            "action": action, "cadence": jours, "raison_skip": raison,
            "billet_tiers": False,
        })
    return entries


def parse_billets_tiers(ws):
    entries = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        nom = cell(r, 0)
        if not nom or not str(nom).strip():
            continue
        if str(nom).strip().upper() in ("ARCHIVES", "ARCHIVE"):
            break  # tout ce qui suit = archives
        action, jours, raison = classify(cell(r, 5), cell(r, 3))
        entries.append({
            "source": "BILLETS TIERS",
            "artiste": str(nom).strip(),
            "ville": "", "salle": "",
            "date_concert": None,
            "mail": txt(cell(r, 1)),
            "commentaire": txt(cell(r, 5)),
            "relance_raw": txt(cell(r, 3)),
            "spectacles": txt(cell(r, 2)),
            "dernier_recu": as_date(cell(r, 4)),
            "action": action, "cadence": jours, "raison_skip": raison,
            "billet_tiers": True,
        })
    return entries


def _fmt(d):
    return d.strftime("%d/%m/%Y") if d else ""


def emails_for_run(entries):
    """Sélectionne les entrées à relancer aujourd'hui et regroupe par destinataire."""
    due = []
    for e in entries:
        if e["action"] != "RELANCE":
            continue
        if JOUR_RUN not in e["cadence"]:
            continue
        if not e["mail"] or "@" not in e["mail"]:
            continue  # pas d'adresse exploitable (souvent un lien/portail)
        due.append(e)
    groups = collections.defaultdict(list)
    for e in due:
        # un destinataire peut avoir plusieurs adresses séparées par espace/slash
        groups[e["mail"].split()[0].strip().lower()].append(e)
    return groups


def render_email(mail, items):
    tiers = any(i["billet_tiers"] for i in items)
    artistes = sorted({i["artiste"] for i in items if i["artiste"]})
    objet = ("[Billet tiers] " if tiers else "") + \
        "Point de vente — " + ", ".join(artistes)
    lignes = []
    for i in items:
        d = _fmt(i["date_concert"])
        loc = " · ".join(x for x in [i["ville"], i["salle"]] if x)
        sp = i.get("spectacles", "")
        desc = " — ".join(x for x in [i["artiste"], loc or sp, d] if x)
        last = f" (dernier point reçu : {_fmt(i['dernier_recu'])})" if i["dernier_recu"] else ""
        lignes.append(f"  • {desc}{last}")
    corps = f"""Objet : {objet}
À : {mail}

Bonjour,

Dans le cadre du suivi des ventes, pourriez-vous nous transmettre le point
de vente à jour pour :

""" + "\n".join(lignes) + """

Merci de répondre directement à ce mail (pointage@bleucitron.net) avec les
chiffres à jour.

Bien à vous,
L'équipe Bleu Citron
"""
    if tiers:
        corps += "\n[Billet tiers — à reporter manuellement dans Aparté/PIMS]\n"
    return corps


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    entries = []
    entries += parse_relances(wb["RELANCES  PIMS A SAISIR"], "RELANCES")
    entries += parse_relances(wb["AUTONOMES  PIMS A SAISIR"], "AUTONOMES")
    entries += parse_billets_tiers(wb["BILLETS TIERS"])

    by_action = collections.Counter(e["action"] for e in entries)
    skip_reasons = collections.Counter(
        e["raison_skip"] for e in entries if e["action"] == "SKIP")

    print(f"RUN DATE : {RUN_DATE} ({JOUR_RUN})")
    print(f"Total entrées normalisées : {len(entries)}")
    print("Répartition action :", dict(by_action))
    print("Raisons de skip :", dict(skip_reasons.most_common()))

    groups = emails_for_run(entries)
    n_mails = len(groups)
    n_lignes = sum(len(v) for v in groups.values())
    print(f"\n>>> LOT DU {JOUR_RUN} : {n_mails} mails individuels "
          f"({n_lignes} pointages demandés)\n")

    for k, items in list(groups.items())[:2]:
        print("-" * 70)
        print(render_email(items[0]["mail"], items))


if __name__ == "__main__":
    main()
